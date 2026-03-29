"""
sources/powerbi.py
------------------
Power BI AR Report integration module.

Two sets of functionality:
  A) AR Report download + cross-reference (new — browser-based export)
     - download_powerbi_ar_report()  → downloads Due & Under Payment AR data
     - load_ar_claims()              → reads Excel, filters Total Outstanding > 0
     - is_claim_in_ar()              → checks if a Claim.MD denial is in the AR
     - get_ar_work_queue()           → sorted work queue from AR data
  B) Legacy Power BI client (API + browser scrape for Billing Summary Notes)
     - PowerBIClient.get_outstanding_claims()
     - powerbi_row_to_claim()
     - parse_billing_summary_csv()

Power BI Report:
  Workspace: 8d724e00-8c1d-4d3c-b804-86c163a258c5
  Report ID:  39dcf41c-1d1b-428a-9086-a8c7e1f3c0f8
"""
from __future__ import annotations

import asyncio
import csv
import glob as glob_mod
import os
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import aiohttp

from config.models import Claim, ClaimStatus, DenialCode, MCO, Program
from config.settings import (
    AZURE_CLIENT_ID,
    AZURE_CLIENT_SECRET,
    AZURE_TENANT_ID,
    POWERBI_REPORT_ID,
    POWERBI_WORKSPACE_ID,
    SESSION_DIR,
    DRY_RUN,
)
from sources.browser_base import BrowserSession
from logging_utils.logger import get_logger

logger = get_logger("powerbi")

# ---------------------------------------------------------------------------
# URLs and constants
# ---------------------------------------------------------------------------

POWERBI_URL = (
    f"https://app.powerbi.com/groups/{POWERBI_WORKSPACE_ID}"
    f"/reports/{POWERBI_REPORT_ID}"
    f"/fb7ccef515505726ae75?experience=power-bi"
)

POWERBI_REPORT_URL = os.getenv(
    "POWERBI_REPORT_URL",
    POWERBI_URL,
)
POWERBI_EMAIL = os.getenv("POWERBI_EMAIL", "nm@lifeconsultantsinc.org")
POWERBI_PASSWORD = os.getenv("POWERBI_PASSWORD", "")

TOKEN_URL   = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
PBI_API_BASE = "https://api.powerbi.com/v1.0/myorg"

# AR report save location
AR_REPORT_DIR = Path("data")
AR_REPORT_DIR.mkdir(parents=True, exist_ok=True)

# Column aliases — matches Power BI report column names (update to actual headers)
COLUMN_MAP = {
    "claim_id":      ["Claim ID", "ClaimID", "Claim #", "claim_id"],
    "client_name":   ["Client Name", "Patient Name", "Member Name"],
    "client_id":     ["Member ID", "Client ID", "Insured ID"],
    "dos":           ["DOS", "Date of Service", "Service Date"],
    "mco":           ["MCO", "Payer", "Insurance"],
    "program":       ["Program", "Billing Region", "Company"],
    "billed_amount": ["Billed Amount", "Charge", "Amount Billed"],
    "paid_amount":   ["Paid Amount", "Payment", "Amount Paid"],
    "status":        ["Status", "Claim Status"],
    "denial_reason": ["Denial Reason", "Rejection Reason", "Notes"],
    "date_billed":   ["Billed Date", "Submit Date", "Submission Date"],
    "auth_number":   ["Auth #", "Authorization", "Auth Number"],
    "note":          ["Note", "Claim Note", "Last Note"],
}

# AR Report Excel columns (as exported from Power BI middle table)
AR_COLUMN_MAP = {
    "consumer_name":     ["Consumer Name", "consumer_name"],
    "unique_id":         ["Unique ID", "unique_id"],
    "service_name":      ["Service Name", "service_name"],
    "document_date":     ["Document Date", "document_date"],
    "billing_units":     ["Billing Units", "billing_units"],
    "billing_amount":    ["Billing Amount", "billing_amount"],
    "total_outstanding":  ["Total Outstanding", "total_outstanding"],
    "first_invoice_date": ["First Invoice Date", "first_invoice_date"],
    "member_number":     ["Member #", "Member Number", "member_number"],
    "mco":               ["MCO", "mco"],
    "deposit_date":      ["Deposit Date", "deposit_date"],
    "total_received":    ["Total Received", "total_received"],
    "check_number":      ["Check Number", "check_number"],
    "write_off_amount":  ["Write Off Amount", "write_off_amount"],
}


# =====================================================================
# PART A: AR Report Download + Cross-Reference (new functionality)
# =====================================================================

class PowerBIARSession(BrowserSession):
    """Browser session for downloading the AR report from Power BI."""

    SESSION_NAME = "powerbi"

    @property
    def login_url(self) -> str:
        return "https://app.powerbi.com/"

    async def _is_logged_in(self) -> bool:
        try:
            # Check for Power BI user avatar or workspace indicators
            el = await self.page.query_selector(
                ".user-picture, .identity-banner, [data-testid='userPhoto'], "
                "[class*='pbiAppHeaderRight'], [class*='avatar']"
            )
            return el is not None
        except Exception:
            return False

    async def _perform_login(self) -> bool:
        """Login to Power BI via Microsoft SSO."""
        await self.page.goto("https://app.powerbi.com/", wait_until="domcontentloaded")
        await asyncio.sleep(3)

        if await self._is_logged_in():
            return True

        try:
            # Microsoft SSO login flow
            # Step 1: Enter email
            email_field = await self.page.wait_for_selector(
                "input[type='email'], input[name='loginfmt']",
                timeout=10_000,
            )
            if email_field:
                await email_field.fill(POWERBI_EMAIL)
                await asyncio.sleep(0.5)
                # Click Next
                next_btn = await self.page.query_selector(
                    "input[type='submit'], button[type='submit'], #idSIButton9"
                )
                if next_btn:
                    await next_btn.click()
                else:
                    await self.page.keyboard.press("Enter")
                await asyncio.sleep(3)

            # Step 2: Enter password
            pwd_field = await self.page.wait_for_selector(
                "input[type='password'], input[name='passwd']",
                timeout=10_000,
            )
            if pwd_field:
                await pwd_field.fill(POWERBI_PASSWORD)
                await asyncio.sleep(0.5)
                sign_in_btn = await self.page.query_selector(
                    "input[type='submit'], button[type='submit'], #idSIButton9"
                )
                if sign_in_btn:
                    await sign_in_btn.click()
                else:
                    await self.page.keyboard.press("Enter")
                await asyncio.sleep(3)

            # Step 3: "Stay signed in?" prompt — click Yes
            try:
                stay_btn = await self.page.wait_for_selector(
                    "#idSIButton9, button:has-text('Yes')",
                    timeout=5_000,
                )
                if stay_btn:
                    await stay_btn.click()
                    await asyncio.sleep(2)
            except Exception:
                pass  # No "Stay signed in" prompt

            # Verify login
            await asyncio.sleep(3)
            return await self._is_logged_in()

        except Exception as e:
            self.logger.warning("Power BI SSO login failed", error=str(e))
            await self.screenshot("powerbi_login_error")
            return await self.handle_mfa("manual")

    async def download_ar_report(self) -> str:
        """
        Navigate to the Power BI report, filter to Due & Under Payment,
        export the middle table, and return the downloaded file path.
        """
        output_path = str(AR_REPORT_DIR / f"powerbi_ar_report.xlsx")

        logger.info("Navigating to Power BI AR report", url=POWERBI_REPORT_URL)
        await self.page.goto(POWERBI_REPORT_URL, wait_until="domcontentloaded", timeout=60_000)

        # Step 1: Wait for report to fully load (visuals, slicers, etc.)
        logger.info("Waiting 20 seconds for report to fully load")
        await asyncio.sleep(20)
        await self.screenshot("powerbi_loaded")

        # Step 2: Set date range — 1 year lookback
        one_year_ago = (date.today() - timedelta(days=365)).strftime("%m/%d/%Y")
        today_str = date.today().strftime("%m/%d/%Y")

        logger.info("Setting date range", start=one_year_ago, end=today_str)
        try:
            await self._set_date_range(one_year_ago, today_str)
        except Exception as e:
            logger.warning("Date range setting failed, continuing with defaults", error=str(e))

        # Step 3: Dismiss any overlay and select AR Status slicers
        logger.info("Selecting Due and Under Payment AR status filters")
        await self._select_ar_status_filters()

        # Step 4: Wait for data to refresh
        logger.info("Waiting 5 seconds for filtered data to load")
        await asyncio.sleep(5)
        await self.screenshot("powerbi_filtered")

        # Step 5: Click into the middle table to activate it
        logger.info("Activating middle table for export")
        await self._activate_middle_table()

        # Step 6: Export data via three-dots menu
        logger.info("Exporting data from middle table")
        downloaded_path = await self._export_table_data()

        if downloaded_path:
            # Move/copy to our standard location
            import shutil
            shutil.copy2(downloaded_path, output_path)
            logger.info("AR report saved", path=output_path, source=downloaded_path)
            return output_path
        else:
            logger.error("AR report download failed — no file received")
            await self.screenshot("powerbi_export_failed")
            raise RuntimeError("Power BI AR report export failed — no download received")

    async def _set_date_range(self, start_date: str, end_date: str):
        """Set the date range inputs on the report."""
        # The date inputs are at approximate positions (290,141) and (370,141)
        # Try selector-based approach first, fall back to coordinates
        try:
            date_inputs = await self.page.query_selector_all(
                "input[type='text'][class*='date'], input[aria-label*='date'], "
                "input[class*='slicer'], input.inputbox"
            )
            if len(date_inputs) >= 2:
                await date_inputs[0].triple_click()
                await date_inputs[0].fill(start_date)
                await self.page.keyboard.press("Tab")
                await asyncio.sleep(0.5)
                await date_inputs[1].triple_click()
                await date_inputs[1].fill(end_date)
                await self.page.keyboard.press("Enter")
                await asyncio.sleep(2)
                return
        except Exception:
            pass

        # Fallback: use mouse coordinates
        logger.info("Using coordinate-based date input (fallback)")

        # Start date input
        await self.page.mouse.click(290, 141)
        await asyncio.sleep(0.5)
        await self.page.keyboard.press("Control+A")
        await self.page.keyboard.type(start_date, delay=50)
        await self.page.keyboard.press("Tab")
        await asyncio.sleep(0.5)

        # End date input
        await self.page.mouse.click(370, 141)
        await asyncio.sleep(0.5)
        await self.page.keyboard.press("Control+A")
        await self.page.keyboard.type(end_date, delay=50)
        await self.page.keyboard.press("Enter")
        await asyncio.sleep(2)

    async def _select_ar_status_filters(self):
        """Select 'Due' and 'Under Payment' in the AR Status slicer."""
        # Click neutral area first to deselect anything
        await self.page.mouse.click(700, 100)
        await asyncio.sleep(0.5)

        # Press Escape to dismiss any overlay
        await self.page.keyboard.press("Escape")
        await asyncio.sleep(0.5)

        # Try to find slicer items by text first
        try:
            due_item = await self.page.query_selector(
                "span:has-text('Due'), div:has-text('Due')"
            )
            under_payment_item = await self.page.query_selector(
                "span:has-text('Under Payment'), div:has-text('Under Payment')"
            )
            if due_item and under_payment_item:
                # Use Meta key (Cmd on Mac) for multi-select
                await due_item.click(modifiers=["Meta"])
                await asyncio.sleep(0.5)
                await under_payment_item.click(modifiers=["Meta"])
                await asyncio.sleep(1)
                return
        except Exception:
            pass

        # Fallback: use mouse coordinates from live testing
        logger.info("Using coordinate-based slicer selection (fallback)")

        # Cmd+click "Due" at approximate position (277, 470)
        await self.page.mouse.click(277, 470, modifiers=["Meta"])
        await asyncio.sleep(0.5)

        # Cmd+click "Under Payment" at approximate position (277, 484)
        await self.page.mouse.click(277, 484, modifiers=["Meta"])
        await asyncio.sleep(1)

    async def _activate_middle_table(self):
        """Click into the middle table to make it the active visual for export."""
        # Click a row in the middle table area (~500, 340)
        await self.page.mouse.click(500, 340)
        await asyncio.sleep(1.5)
        # Click again to ensure activation
        await self.page.mouse.click(500, 340)
        await asyncio.sleep(1)

    async def _export_table_data(self) -> Optional[str]:
        """
        Export data from the active table visual using the three-dots menu.
        Returns the path of the downloaded file, or None on failure.
        """
        # Step 1: Find the three-dots button for the middle table
        # Look for button.vcMenuBtn with y-position between 250-290
        three_dots = await self._find_three_dots_button()

        if not three_dots:
            logger.warning("Three-dots menu button not found, retrying with coordinate click")
            # Fallback: try right-click on the table
            await self.page.mouse.click(500, 340, button="right")
            await asyncio.sleep(1)
        else:
            # Click the three-dots button using its coordinates
            box = await three_dots.bounding_box()
            if box:
                await self.page.mouse.click(
                    box["x"] + box["width"] / 2,
                    box["y"] + box["height"] / 2,
                )
            else:
                await three_dots.click()
            await asyncio.sleep(1)

        # Step 2: Click "Export data" from the dropdown
        try:
            export_option = await self.page.wait_for_selector(
                "button:has-text('Export data'), "
                "span:has-text('Export data'), "
                "div[role='menuitem']:has-text('Export data'), "
                "[class*='menu'] :has-text('Export data')",
                timeout=5_000,
            )
            if export_option:
                await export_option.click()
                await asyncio.sleep(3)
            else:
                raise Exception("Export data option not found in menu")
        except Exception as e:
            logger.warning("Could not find Export data menu item", error=str(e))
            await self.screenshot("powerbi_no_export_menu")
            # Try typing it
            await self.page.keyboard.type("Export data")
            await self.page.keyboard.press("Enter")
            await asyncio.sleep(3)

        # Step 3: Click the green "Export" button in the dialog
        try:
            # Find the Export button (highest Y position = most likely the confirm button)
            export_buttons = await self.page.query_selector_all(
                "button:has-text('Export')"
            )
            target_btn = None
            max_y = 0
            for btn in export_buttons:
                box = await btn.bounding_box()
                if box and box["y"] > max_y:
                    max_y = box["y"]
                    target_btn = btn

            if target_btn:
                # Set up download listener before clicking
                async with self.page.expect_download(timeout=60_000) as dl_info:
                    await target_btn.click()
                download = await dl_info.value

                # Save the downloaded file
                dl_path = str(
                    AR_REPORT_DIR
                    / f"powerbi_ar_raw_{date.today().isoformat()}.xlsx"
                )
                await download.save_as(dl_path)
                logger.info("Download complete", path=dl_path)
                return dl_path
            else:
                raise Exception("Export confirmation button not found")
        except Exception as e:
            logger.error("Export button click or download failed", error=str(e))
            await self.screenshot("powerbi_export_dialog_error")
            return None

    async def _find_three_dots_button(self):
        """
        Find the three-dots (vcMenuBtn) button for the middle table.
        The correct one has a y-position between 250 and 290.
        """
        try:
            buttons = await self.page.query_selector_all("button.vcMenuBtn")
            for btn in buttons:
                box = await btn.bounding_box()
                if box and 250 <= box["y"] <= 290:
                    return btn

            # Broader search if exact class not found
            buttons = await self.page.query_selector_all(
                "button[class*='menu'], button[class*='Menu'], "
                "button[aria-label*='More options'], button[title*='More options']"
            )
            for btn in buttons:
                box = await btn.bounding_box()
                if box and 250 <= box["y"] <= 290:
                    return btn
        except Exception as e:
            logger.warning("Three-dots button search failed", error=str(e))

        return None


# ---------------------------------------------------------------------------
# Public AR Report API
# ---------------------------------------------------------------------------

async def download_powerbi_ar_report() -> str:
    """
    Download the AR report from Power BI.

    Logs into Power BI (reusing saved session when possible),
    sets date filter to 1-year lookback,
    filters to Due & Under Payment AR statuses,
    exports the middle table data,
    and saves to data/powerbi_ar_report.xlsx.

    Returns the file path of the saved report.
    """
    if DRY_RUN:
        logger.info("DRY_RUN: Skipping Power BI AR report download")
        # Check if we have a recent report to use
        existing = str(AR_REPORT_DIR / "powerbi_ar_report.xlsx")
        if Path(existing).exists():
            logger.info("DRY_RUN: Using existing AR report", path=existing)
            return existing
        raise RuntimeError("DRY_RUN: No existing AR report found at data/powerbi_ar_report.xlsx")

    async with PowerBIARSession(headless=True) as session:
        return await session.download_ar_report()


def load_ar_claims(xlsx_path: str) -> List[Dict]:
    """
    Read the downloaded AR report Excel file.
    Filters to rows where Total Outstanding > 0.

    Returns list of dicts with normalized keys:
        consumer_name, unique_id, service_name, document_date,
        billing_amount, total_outstanding, member_number, mco,
        first_invoice_date, total_received

    Uses openpyxl read_only mode for performance (17K+ rows).
    """
    from openpyxl import load_workbook

    logger.info("Loading AR report", path=xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        logger.error("AR report is empty — no header row")
        wb.close()
        return []

    headers = [str(h).strip() if h else "" for h in raw_headers]

    # Build column index mapping
    def _find_col(keys: List[str]) -> Optional[int]:
        for key in keys:
            for i, h in enumerate(headers):
                if h.lower() == key.lower():
                    return i
        return None

    col_map = {}
    for field_name, aliases in AR_COLUMN_MAP.items():
        idx = _find_col(aliases)
        col_map[field_name] = idx

    # Read data rows
    claims = []
    total_rows = 0
    filtered_out = 0

    for row in rows_iter:
        total_rows += 1

        def _get_val(field_name: str):
            idx = col_map.get(field_name)
            if idx is not None and idx < len(row):
                val = row[idx]
                return val if val is not None else ""
            return ""

        # Parse Total Outstanding
        outstanding_raw = _get_val("total_outstanding")
        try:
            total_outstanding = float(
                re.sub(r"[^\d.\-]", "", str(outstanding_raw))
            ) if outstanding_raw else 0.0
        except (ValueError, TypeError):
            total_outstanding = 0.0

        # Filter: only keep rows with Total Outstanding > 0
        if total_outstanding <= 0:
            filtered_out += 1
            continue

        # Parse billing amount
        billing_raw = _get_val("billing_amount")
        try:
            billing_amount = float(
                re.sub(r"[^\d.\-]", "", str(billing_raw))
            ) if billing_raw else 0.0
        except (ValueError, TypeError):
            billing_amount = 0.0

        # Parse total received
        received_raw = _get_val("total_received")
        try:
            total_received = float(
                re.sub(r"[^\d.\-]", "", str(received_raw))
            ) if received_raw else 0.0
        except (ValueError, TypeError):
            total_received = 0.0

        # Parse dates
        doc_date = _parse_date(str(_get_val("document_date")))
        first_inv = _parse_date(str(_get_val("first_invoice_date")))

        # Parse member number (strip whitespace, keep leading zeros)
        member_num = str(_get_val("member_number")).strip()

        claim = {
            "consumer_name": str(_get_val("consumer_name")).strip(),
            "unique_id": str(_get_val("unique_id")).strip(),
            "service_name": str(_get_val("service_name")).strip(),
            "document_date": doc_date,
            "billing_amount": billing_amount,
            "total_outstanding": total_outstanding,
            "member_number": member_num,
            "mco": str(_get_val("mco")).strip(),
            "first_invoice_date": first_inv,
            "total_received": total_received,
        }
        claims.append(claim)

    wb.close()

    logger.info(
        "AR report loaded",
        total_rows=total_rows,
        outstanding_claims=len(claims),
        filtered_out=filtered_out,
    )
    return claims


def is_claim_in_ar(
    member_number: str,
    dos: date,
    ar_claims: List[Dict],
) -> Optional[Dict]:
    """
    Check if a specific claim exists in the AR data by Member # and DOS.

    Args:
        member_number: The Medicaid/Member number (from Claim.MD client_id)
        dos: Date of service (from Claim.MD dos field)
        ar_claims: The loaded AR claims list from load_ar_claims()

    Returns:
        The matching AR record dict, or None if not found.
    """
    if not member_number or not dos or not ar_claims:
        return None

    # Normalize the member number for comparison (strip, lowercase)
    norm_member = member_number.strip().lower()

    for ar_claim in ar_claims:
        ar_member = str(ar_claim.get("member_number", "")).strip().lower()
        ar_dos = ar_claim.get("document_date")

        if ar_member == norm_member and ar_dos == dos:
            return ar_claim

    return None


def get_ar_work_queue(ar_claims: List[Dict]) -> List[Dict]:
    """
    Return the filtered list of claims that need work,
    sorted by Total Outstanding descending (highest value first).

    Args:
        ar_claims: The loaded AR claims list from load_ar_claims()
                   (already filtered to Total Outstanding > 0)

    Returns:
        Sorted list of AR claim dicts.
    """
    # Already filtered by load_ar_claims, just sort
    sorted_claims = sorted(
        ar_claims,
        key=lambda c: c.get("total_outstanding", 0),
        reverse=True,
    )

    logger.info(
        "AR work queue prepared",
        total_claims=len(sorted_claims),
        total_outstanding=sum(c.get("total_outstanding", 0) for c in sorted_claims),
    )
    return sorted_claims


# =====================================================================
# PART B: Legacy Power BI Client (Azure AD API + browser scrape)
# =====================================================================

class AzureTokenManager:
    def __init__(self):
        self._token: Optional[str] = None
        self._expires_at: float = 0.0

    async def get_token(self) -> Optional[str]:
        import time
        if self._token and time.time() < self._expires_at - 60:
            return self._token
        if not all([AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET]):
            return None
        url = TOKEN_URL.format(tenant=AZURE_TENANT_ID)
        data = {
            "grant_type":    "client_credentials",
            "client_id":     AZURE_CLIENT_ID,
            "client_secret": AZURE_CLIENT_SECRET,
            "scope":         "https://analysis.windows.net/powerbi/api/.default",
        }
        async with aiohttp.ClientSession() as session:
            async with session.post(url, data=data) as resp:
                if resp.status == 200:
                    body = await resp.json()
                    self._token = body["access_token"]
                    self._expires_at = time.time() + body.get("expires_in", 3600)
                    return self._token
        logger.error("Azure AD token acquisition failed")
        return None


_token_mgr = AzureTokenManager()


class PowerBIClient:

    async def get_outstanding_claims(self) -> List[Claim]:
        token = await _token_mgr.get_token()
        if token:
            try:
                return await self._fetch_via_api(token)
            except Exception as e:
                logger.warning("Power BI API failed, falling back to browser", error=str(e))
        return await self._fetch_via_browser()

    # -- API path --

    async def _fetch_via_api(self, token: str) -> List[Claim]:
        dataset_id = await self._get_dataset_id(token)
        if not dataset_id:
            raise RuntimeError("Dataset ID not found")

        dax = """
        EVALUATE
        FILTER(
            'BillingSummaryNotes',
            'BillingSummaryNotes'[Status] <> "Paid"
            && 'BillingSummaryNotes'[Status] <> "Written Off"
        )
        ORDER BY 'BillingSummaryNotes'[DOS] ASC
        """
        headers = {"Authorization": f"Bearer {token}"}
        url = f"{PBI_API_BASE}/datasets/{dataset_id}/executeQueries"
        payload = {"queries": [{"query": dax}], "serializerSettings": {"includeNulls": True}}

        async with aiohttp.ClientSession() as session:
            async with session.post(url, json=payload, headers=headers) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    rows = (data.get("results", [{}])[0]
                               .get("tables", [{}])[0]
                               .get("rows", []))
                    logger.info("Power BI API rows", count=len(rows))
                    return [self._row_to_claim(r) for r in rows if r]
                raise RuntimeError(f"DAX query {resp.status}: {await resp.text()}")

    async def _get_dataset_id(self, token: str) -> Optional[str]:
        url = f"{PBI_API_BASE}/groups/{POWERBI_WORKSPACE_ID}/reports/{POWERBI_REPORT_ID}"
        headers = {"Authorization": f"Bearer {token}"}
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as resp:
                if resp.status == 200:
                    return (await resp.json()).get("datasetId")
        return None

    # -- Browser fallback --

    async def _fetch_via_browser(self) -> List[Claim]:
        logger.info("Fetching Power BI data via browser")
        async with PowerBIBrowserSession() as session:
            return await session.scrape_outstanding_claims()

    # -- Row parser --

    def _row_to_claim(self, row: dict) -> Claim:
        def get(keys):
            for k in keys:
                if k in row:
                    return str(row[k]).strip()
                for rk in row:
                    if k.lower() in rk.lower():
                        return str(row[rk]).strip()
            return ""

        dos         = _parse_date(get(COLUMN_MAP["dos"]))
        date_billed = _parse_date(get(COLUMN_MAP["date_billed"]))
        billed      = _parse_float(get(COLUMN_MAP["billed_amount"]))
        paid        = _parse_float(get(COLUMN_MAP["paid_amount"]))
        mco_str     = get(COLUMN_MAP["mco"])
        program_str = get(COLUMN_MAP["program"])
        denial_raw  = get(COLUMN_MAP["denial_reason"])
        age_days    = (date.today() - date_billed).days if date_billed else 0

        return Claim(
            claim_id      = get(COLUMN_MAP["claim_id"]) or f"PBI_{id(row)}",
            client_name   = get(COLUMN_MAP["client_name"]),
            client_id     = get(COLUMN_MAP["client_id"]),
            dos           = dos or date.today(),
            mco           = _parse_mco(mco_str),
            program       = _parse_program(program_str),
            billed_amount = billed,
            paid_amount   = paid,
            status        = _parse_status(get(COLUMN_MAP["status"])),
            denial_codes  = _parse_denial_codes_text(denial_raw),
            denial_reason_raw = denial_raw,
            date_billed   = date_billed,
            auth_number   = get(COLUMN_MAP["auth_number"]),
            last_note     = get(COLUMN_MAP["note"]),
            age_days      = age_days,
        )


# ---------------------------------------------------------------------------
# Browser session scraper (legacy)
# ---------------------------------------------------------------------------

class PowerBIBrowserSession(BrowserSession):
    SESSION_NAME = "powerbi"

    @property
    def login_url(self) -> str:
        return "https://app.powerbi.com/"

    async def _is_logged_in(self) -> bool:
        try:
            el = await self.page.query_selector(
                ".user-picture, .identity-banner, [data-testid='userPhoto']"
            )
            return el is not None
        except Exception:
            return False

    async def _perform_login(self) -> bool:
        await self.page.goto(self.login_url)
        await asyncio.sleep(2)
        if await self._is_logged_in():
            return True
        logger.warning("Power BI requires Microsoft SSO — manual login needed")
        return await self.handle_mfa("manual")

    async def scrape_outstanding_claims(self) -> List[Claim]:
        claims = []
        try:
            await self.page.goto(POWERBI_URL, wait_until="networkidle", timeout=60_000)
            await asyncio.sleep(5)

            # Locate the table visual
            table = await self.page.query_selector(
                ".tableEx, [aria-label*='claim'], .visual-container table, .pivotTable"
            )
            if table:
                claims = await self._parse_pbi_table(table)
            else:
                claims = await self._export_csv()
        except Exception as e:
            logger.error("Power BI scrape failed", error=str(e))
            await self.screenshot("powerbi_error")
        return claims

    async def _parse_pbi_table(self, table) -> List[Claim]:
        headers = []
        header_cells = await table.query_selector_all(
            "th, .columnHeader, [role='columnheader']"
        )
        for cell in header_cells:
            headers.append((await cell.inner_text()).strip())

        claims = []
        client = PowerBIClient()
        rows = await table.query_selector_all("tr[role='row'], tbody tr")
        for row in rows:
            cells = await row.query_selector_all("td, [role='cell']")
            if not cells:
                continue
            row_data = {
                headers[i]: (await cell.inner_text()).strip()
                for i, cell in enumerate(cells)
                if i < len(headers)
            }
            if row_data:
                claims.append(client._row_to_claim(row_data))
        return claims

    async def _export_csv(self) -> List[Claim]:
        claims = []
        try:
            visual = await self.page.query_selector(".visual-container")
            if visual:
                await visual.click(button="right")
                await asyncio.sleep(0.5)
                export_opt = await self.page.query_selector("text='Export data'")
                if export_opt:
                    await export_opt.click()
                    await asyncio.sleep(1)
                    async with self.page.expect_download(timeout=30_000) as dl_info:
                        btn = await self.page.query_selector("button:has-text('Export')")
                        if btn:
                            await btn.click()
                    dl = await dl_info.value
                    path = f"/tmp/claims_work/pbi_{date.today().isoformat()}.csv"
                    await dl.save_as(path)
                    claims = _parse_csv(path)
        except Exception as e:
            logger.error("PBI CSV export failed", error=str(e))
        return claims


def _parse_csv(path: str) -> List[Claim]:
    client = PowerBIClient()
    claims = []
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                claims.append(client._row_to_claim(dict(row)))
    except Exception as e:
        logger.error("CSV parse error", error=str(e))
    return claims


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _parse_date(s: str) -> Optional[date]:
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            pass
    # Handle Excel serial date numbers
    try:
        serial = float(s.strip())
        if 30000 < serial < 60000:
            from datetime import timedelta as td
            return date(1899, 12, 30) + td(days=int(serial))
    except (ValueError, TypeError):
        pass
    return None


def _parse_float(s: str) -> float:
    try:
        return float(re.sub(r"[^\d.]", "", s))
    except (ValueError, TypeError):
        return 0.0


def _parse_mco(s: str) -> MCO:
    u = s.upper()
    if "UNITED" in u or "UHC" in u: return MCO.UNITED
    if "SENTARA" in u:              return MCO.SENTARA
    if "AETNA" in u:                return MCO.AETNA
    if "ANTHEM" in u:               return MCO.ANTHEM
    if "MOLINA" in u:               return MCO.MOLINA
    if "HUMANA" in u:               return MCO.HUMANA
    if "MAGELLAN" in u:             return MCO.MAGELLAN
    if "MEDICAID" in u or "DMAS" in u: return MCO.DMAS
    return MCO.UNKNOWN


def _parse_program(s: str) -> Program:
    u = s.upper()
    if "KJLN" in u:                    return Program.KJLN
    if "NHCS" in u or "NEW HEIGHTS" in u: return Program.NHCS
    if "MARY" in u:                    return Program.MARYS_HOME
    return Program.UNKNOWN


def _parse_status(s: str) -> ClaimStatus:
    l = s.lower()
    if "paid" in l:      return ClaimStatus.PAID
    if "denied" in l:    return ClaimStatus.DENIED
    if "rejected" in l:  return ClaimStatus.REJECTED
    if "recon" in l:     return ClaimStatus.IN_RECON
    if "appeal" in l:    return ClaimStatus.IN_APPEAL
    if "write" in l:     return ClaimStatus.WRITTEN_OFF
    return ClaimStatus.PENDING


def _parse_denial_codes_text(s: str) -> List[DenialCode]:
    from sources.claimmd import parse_denial_codes
    return parse_denial_codes(s)


# ---------------------------------------------------------------------------
# Compatibility shims for test suite
# ---------------------------------------------------------------------------

def _parse_int(s: str) -> int:
    try:
        return int(re.sub(r"[^\d]", "", s))
    except (ValueError, TypeError):
        return 0


def parse_billing_summary_csv(csv_text: str) -> List[Claim]:
    """
    Parse a CSV export of the Power BI Billing Summary Notes report.
    Accepts raw CSV string, returns list of Claims.
    Expected columns: Client Name, MCO, DOS, Amount Billed, Amount Paid,
                      Denial Reason, Billing Program, Auth Number, Days Outstanding
    """
    import csv as csv_mod
    import io
    client = PowerBIClient()
    claims = []
    reader = csv_mod.DictReader(io.StringIO(csv_text))
    for row in reader:
        claims.append(client._row_to_claim(dict(row)))
    return claims


# ---------------------------------------------------------------------------
# Public API used by tests and orchestrator
# ---------------------------------------------------------------------------

def _str_to_mco(s: str) -> MCO:
    """Public alias for _parse_mco."""
    return _parse_mco(s)


def _str_to_program(s: str) -> Program:
    """Public alias for _parse_program."""
    return _parse_program(s)


# Exact column names as exported from Power BI (camelCase + spaced variants)
_EXACT_COL_MAP = {
    "claim_id":      ["ClaimID", "Claim ID", "ClaimId", "claim_id"],
    "client_name":   ["ClientName", "Client Name", "PatientName", "client_name"],
    "client_id":     ["ClientID", "Client ID", "MemberID", "Member ID", "client_id"],
    "dos":           ["DOS", "DateOfService", "Date of Service", "dos"],
    "mco":           ["MCO", "Payer", "Insurance", "mco"],
    "program":       ["BillingRegion", "Billing Region", "Program", "Company", "billing_region"],
    "billed_amount": ["BilledAmount", "Billed Amount", "Charge", "billed_amount"],
    "paid_amount":   ["PaidAmount", "Paid Amount", "Payment", "paid_amount"],
    "status":        ["Status", "ClaimStatus", "Claim Status", "status"],
    "denial_reason": ["DenialReason", "Denial Reason", "RejectionReason", "denial_reason"],
    "date_billed":   ["DateBilled", "SubmitDate", "Submit Date", "Billed Date", "date_billed"],
    "auth_number":   ["AuthNumber", "Auth Number", "Auth #", "Authorization", "auth_number"],
    "note":          ["LastNote", "Note", "Claim Note", "last_note"],
    "days_outstanding": ["DaysOutstanding", "Days Outstanding", "AgeDays", "age_days"],
}


def _exact_get(row: dict, keys: list) -> str:
    """Get a value from row using a list of possible key names (exact match first)."""
    for k in keys:
        if k in row:
            return str(row[k]).strip()
    # Fallback: case-insensitive substring match
    for k in keys:
        for rk in row:
            if k.lower() == rk.lower():
                return str(row[rk]).strip()
    return ""


def powerbi_row_to_claim(row: dict) -> Optional["Claim"]:
    """
    Convert a Power BI report row dict to a Claim object.
    Returns None if the row has no usable Claim ID.
    Adds UNDERPAID denial code if paid < 95% of billed.
    """
    claim_id = _exact_get(row, _EXACT_COL_MAP["claim_id"])
    if not claim_id:
        return None  # Skip rows with no claim identifier

    client_name   = _exact_get(row, _EXACT_COL_MAP["client_name"])
    client_id     = _exact_get(row, _EXACT_COL_MAP["client_id"])
    dos_str       = _exact_get(row, _EXACT_COL_MAP["dos"])
    mco_str       = _exact_get(row, _EXACT_COL_MAP["mco"])
    program_str   = _exact_get(row, _EXACT_COL_MAP["program"])
    billed_str    = _exact_get(row, _EXACT_COL_MAP["billed_amount"])
    paid_str      = _exact_get(row, _EXACT_COL_MAP["paid_amount"])
    status_str    = _exact_get(row, _EXACT_COL_MAP["status"])
    denial_raw    = _exact_get(row, _EXACT_COL_MAP["denial_reason"])
    date_billed_s = _exact_get(row, _EXACT_COL_MAP["date_billed"])
    auth_number   = _exact_get(row, _EXACT_COL_MAP["auth_number"])
    note          = _exact_get(row, _EXACT_COL_MAP["note"])
    days_str      = _exact_get(row, _EXACT_COL_MAP["days_outstanding"])

    dos         = _parse_date(dos_str)
    date_billed = _parse_date(date_billed_s)
    billed      = _parse_float(billed_str)
    paid        = _parse_float(paid_str)
    mco         = _parse_mco(mco_str)
    program     = _parse_program(program_str)
    status      = _parse_status(status_str)

    # Use DaysOutstanding from report if available, else compute from DOS
    if days_str:
        age_days = _parse_int(days_str)
    elif date_billed:
        age_days = (date.today() - date_billed).days
    elif dos:
        age_days = (date.today() - dos).days
    else:
        age_days = 0

    denial_codes = _parse_denial_codes_text(denial_raw)

    # Auto-detect underpaid: paid < 95% of billed (and billed > 0, not fully denied)
    if (
        billed > 0
        and paid > 0
        and paid < billed * 0.95
        and DenialCode.UNDERPAID not in denial_codes
    ):
        denial_codes = [DenialCode.UNDERPAID] + [c for c in denial_codes if c != DenialCode.UNKNOWN]

    return Claim(
        claim_id      = claim_id,
        client_name   = client_name,
        client_id     = client_id,
        dos           = dos or date.today(),
        mco           = mco,
        program       = program,
        billed_amount = billed,
        paid_amount   = paid,
        status        = status,
        denial_codes  = denial_codes if denial_codes else [DenialCode.UNKNOWN],
        denial_reason_raw = denial_raw,
        date_billed   = date_billed,
        auth_number   = auth_number,
        last_note     = note,
        age_days      = age_days,
    )
