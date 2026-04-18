"""
sources/lauris_xml.py
---------------------
Fetches billing, AR, and authorization data directly from Lauris XML
streams using HTTP Basic Auth. Replaces Power BI for the master work queue.

XML Endpoints (Basic Auth with Lauris credentials):
  - Billing Summary: billing amounts, DOS, entity, service, auth join ID
  - AR Information: payments received, check numbers, deposits
  - Authorization Info: auth numbers, periods, MCO, member ID, diagnosis

Join key: Billing_Summary_ID (Billing ↔ AR),
          AuthID__x0028_for_joining_x0029_ (Billing ↔ Auth)
"""
from __future__ import annotations

import os
import json
import requests
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from logging_utils.logger import get_logger
from reporting.report_paths import sync_report_file, unique_report_path

logger = get_logger("lauris_xml")

BASE_URL = "https://www12.laurisonline.com"
USERNAME = os.getenv("LAURIS_USERNAME", "")
PASSWORD = os.getenv("LAURIS_PASSWORD", "")

# Lauris XML view URLs (Basic Auth)
BILLING_SUMMARY_XML = (
    f"{BASE_URL}/reports/formsearchdataviewXML.aspx"
    "?viewid=%2buaLca3%2bmmDX5TfvMH%2f25g%3d%3d"
)
AR_INFO_XML = (
    f"{BASE_URL}/reports/formsearchdataviewXML.aspx"
    "?viewid=fwR8FpcbZLiOYvlDtnhz8A%3d%3d"
)
AUTH_INFO_XML = (
    f"{BASE_URL}/reports/formsearchdataviewXML.aspx"
    "?viewid=E1jRUaNGKAxt%2bAa7Ubk1xg%3d%3d"
)
LAURIS_CACHE_DIR = Path("data") / "cache"
LAURIS_CACHE_TTL_HOURS = 24


def _cache_path(cache_key: str) -> Path:
    safe_key = "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in cache_key)
    return LAURIS_CACHE_DIR / f"{safe_key}.json"


def _read_cached_xml(cache_key: str, max_age_hours: int = LAURIS_CACHE_TTL_HOURS) -> str:
    path = _cache_path(cache_key)
    if not path.exists():
        return ""
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        fetched_at = datetime.fromisoformat(payload.get("fetched_at", ""))
        if datetime.now() - fetched_at > timedelta(hours=max_age_hours):
            return ""
        content = str(payload.get("content", "") or "")
        if content.strip():
            logger.info("Using cached Lauris XML", cache_key=cache_key, path=str(path))
            return content
    except Exception as exc:  # noqa: BLE001
        logger.warning("Ignoring invalid Lauris cache file", cache_key=cache_key, error=str(exc))
    return ""


def _write_cached_xml(cache_key: str, content: str) -> None:
    path = _cache_path(cache_key)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "fetched_at": datetime.now().isoformat(),
        "content": content,
    }
    path.write_text(json.dumps(payload), encoding="utf-8")


def fetch_xml_text(
    url: str,
    *,
    cache_key: str,
    timeout: int = 300,
    max_age_hours: int = LAURIS_CACHE_TTL_HOURS,
) -> str:
    """Fetch Lauris XML via Basic Auth and cache the raw response for 24 hours."""
    cached = _read_cached_xml(cache_key, max_age_hours=max_age_hours)
    if cached:
        return cached

    r = requests.get(url, auth=(USERNAME, PASSWORD), timeout=timeout)
    r.raise_for_status()
    content = r.text
    if not content.strip():
        raise ValueError(f"Empty response from {url}")
    _write_cached_xml(cache_key, content)
    logger.info("Cached Lauris XML", cache_key=cache_key, path=str(_cache_path(cache_key)))
    return content


def _fetch_xml(
    url: str,
    timeout: int = 300,
    *,
    cache_key: str,
    max_age_hours: int = LAURIS_CACHE_TTL_HOURS,
) -> ET.Element:
    """Fetch and parse a Lauris XML stream via Basic Auth with 24-hour disk cache."""
    return ET.fromstring(
        fetch_xml_text(
            url,
            cache_key=cache_key,
            timeout=timeout,
            max_age_hours=max_age_hours,
        )
    )


def fetch_outstanding_claims(
    lookback_days: int = 365,
) -> List[dict]:
    """
    Fetch all outstanding claims (Due + Under Payment) by joining
    Billing Summary, AR Information, and Authorization data.

    Returns a list of claim dicts with billing, auth, and AR fields.
    """
    cutoff = (datetime.now() - timedelta(days=lookback_days)).isoformat()

    # 1. Billing Summary
    logger.info("Fetching Lauris Billing Summary XML...")
    bs_root = _fetch_xml(BILLING_SUMMARY_XML, cache_key="lauris_billing_summary")
    billing: Dict[str, dict] = {}
    for row in bs_root.findall(".//Billing_Summary_View"):
        bs_id = (row.findtext("Billing_Summary_ID") or "").strip()
        doc_date = (row.findtext("Document_Date") or "").strip()
        if not bs_id or (doc_date and doc_date < cutoff):
            continue
        billing[bs_id] = {
            "bs_id": bs_id,
            "name": (row.findtext("Name") or "").strip(),
            "key": (row.findtext("Key") or "").strip(),
            "service": (row.findtext("Service_Name") or "").strip(),
            "billing_amount": float(
                row.findtext("Billing_Amount") or "0"
            ),
            "billing_units": (
                row.findtext("Billing_Units") or ""
            ).strip(),
            "doc_date": doc_date[:10],
            "region": (row.findtext("User_Region") or "").strip(),
            "modifier": (
                row.findtext("Billing_Modifier") or ""
            ).strip(),
            "auth_id": (
                row.findtext(
                    "AuthID__x0028_for_joining_x0029_"
                )
                or ""
            ).strip(),
            "invoice_date": (
                row.findtext("First_Invoice_Date") or ""
            ).strip()[:10],
            "stbid": (row.findtext("STBID") or "").strip(),
            "closed_id": (
                row.findtext("Closed_ID") or ""
            ).strip(),
            # Fields populated by joins below
            "member_id": "",
            "mco": "",
            "total_received": 0.0,
            "check_number": "",
            "deposit_date": "",
            "auth_number": "",
            "auth_start": "",
            "auth_end": "",
            "auth_status": "",
            "billing_diagnosis": "",
            "outstanding": 0.0,
            "ar_status": "",
        }
    logger.info("Billing items loaded", count=len(billing))

    # 2. AR Information (payments)
    logger.info("Fetching Lauris AR Information XML...")
    ar_root = _fetch_xml(AR_INFO_XML, cache_key="lauris_ar_information")
    for row in ar_root.findall(".//AR_Information_View"):
        bs_id = (
            row.findtext("Billing_Summary_ID") or ""
        ).strip()
        rcvd = float(row.findtext("Received_Amount") or "0")
        if bs_id in billing:
            billing[bs_id]["total_received"] += rcvd
            chk = (
                row.findtext("Check_Number") or ""
            ).strip()
            dep = (
                row.findtext("Deposit_Date") or ""
            ).strip()[:10]
            if chk:
                billing[bs_id]["check_number"] = chk
            if dep:
                billing[bs_id]["deposit_date"] = dep

    # 3. Authorization Information
    logger.info("Fetching Lauris Authorization XML...")
    auth_root = _fetch_xml(AUTH_INFO_XML, cache_key="lauris_authorization")
    auth_by_join = {}
    for row in auth_root.findall(
        ".//Authorization_Information_View"
    ):
        join_id = (
            row.findtext(
                "AuthID__x0028_for_joining_x0029_"
            )
            or ""
        ).strip()
        if join_id:
            auth_by_join[join_id] = {
                "auth_number": (
                    row.findtext("Authorization_Number") or ""
                ).strip(),
                "start_date": (
                    row.findtext("Start_Date") or ""
                ).strip()[:10],
                "end_date": (
                    row.findtext("End_Date") or ""
                ).strip()[:10],
                "status": (
                    row.findtext("Authorization_Status") or ""
                ).strip(),
                "payor": (
                    row.findtext("Payor") or ""
                ).strip(),
                "member_id": (
                    row.findtext("Insurance_Policy_No") or ""
                ).strip(),
                "diagnosis": (
                    row.findtext("Billing_Diagnosis") or ""
                ).strip(),
            }
    logger.info("Auth lookup built", count=len(auth_by_join))

    # Link auths to billing
    linked = 0
    for item in billing.values():
        aid = item.get("auth_id", "")
        if aid and aid in auth_by_join:
            a = auth_by_join[aid]
            item["auth_number"] = a["auth_number"]
            item["auth_start"] = a["start_date"]
            item["auth_end"] = a["end_date"]
            item["auth_status"] = a["status"]
            item["mco"] = a["payor"]
            item["member_id"] = a["member_id"]
            item["billing_diagnosis"] = a["diagnosis"]
            linked += 1
    logger.info("Auths linked to billing", linked=linked)

    # 4. Classify: Due vs Under Payment
    results = []
    for item in billing.values():
        amt = item["billing_amount"]
        rcvd = item["total_received"]
        if amt <= 0:
            continue
        outstanding = amt - rcvd
        item["outstanding"] = outstanding

        if rcvd == 0:
            item["ar_status"] = "Due"
            results.append(item)
        elif rcvd < amt - 0.01:
            item["ar_status"] = "Under Payment"
            results.append(item)

    due_count = sum(1 for r in results if r["ar_status"] == "Due")
    under_count = len(results) - due_count
    total_outstanding = sum(r["outstanding"] for r in results)
    logger.info(
        "Outstanding claims identified",
        due=due_count,
        under_payment=under_count,
        total=len(results),
        total_outstanding=f"${total_outstanding:,.2f}",
    )

    return results


def fetch_billing_bridge(lookback_days: int = 365) -> list[dict]:
    """
    Fetch the Billing Summary bridge rows used to map Claim.MD claims to Lauris.

    This is lighter than the full outstanding-claims join when code only needs:
      - Billing_Summary_ID
      - STBID / patient account number
      - Key / Unique_ID
      - Name
      - Document_Date
      - AuthID
    """
    cutoff = (datetime.now() - timedelta(days=lookback_days)).isoformat()
    logger.info("Fetching Lauris Billing Summary bridge XML...")
    root = _fetch_xml(BILLING_SUMMARY_XML, cache_key="lauris_billing_summary")
    rows: list[dict] = []
    for row in root.findall(".//Billing_Summary_View"):
        bs_id = (row.findtext("Billing_Summary_ID") or "").strip()
        doc_date = (row.findtext("Document_Date") or "").strip()
        if not bs_id or (doc_date and doc_date < cutoff):
            continue
        rows.append({
            "bs_id": bs_id,
            "patient_account_number": (row.findtext("STBID") or "").strip(),
            "lauris_id": (row.findtext("Key") or "").strip(),
            "name": (row.findtext("Name") or "").strip(),
            "auth_id": (
                row.findtext("AuthID__x0028_for_joining_x0029_") or ""
            ).strip(),
            "doc_date": doc_date[:10],
        })
    logger.info("Billing bridge rows loaded", count=len(rows))
    return rows


def fetch_claim_member_bridge(lookback_days: int = 365) -> list[dict]:
    """
    Fetch Lauris claim bridge rows keyed the same way Claim.MD claims are matched.

    The reliable join from Claim.MD into Lauris is:
      Claim.client_id / ins_number  -> Lauris auth Insurance_Policy_No
      Claim.dos                     -> Lauris Billing Summary Document_Date

    This helper returns joined billing+auth rows with the pieces needed to
    resolve demographics and authorization details before payer API calls.
    """
    cutoff = (datetime.now() - timedelta(days=lookback_days)).isoformat()

    logger.info("Fetching Lauris Billing Summary XML for claim member bridge...")
    billing_root = _fetch_xml(BILLING_SUMMARY_XML, cache_key="lauris_billing_summary")
    logger.info("Fetching Lauris Authorization XML for claim member bridge...")
    auth_root = _fetch_xml(AUTH_INFO_XML, cache_key="lauris_authorization")

    auth_by_join: dict[str, dict[str, str]] = {}
    for row in auth_root.findall(".//Authorization_Information_View"):
        join_id = (
            row.findtext("AuthID__x0028_for_joining_x0029_")
            or ""
        ).strip()
        if not join_id:
            continue
        auth_by_join[join_id] = {
            "auth_number": (row.findtext("Authorization_Number") or "").strip(),
            "auth_status": (row.findtext("Authorization_Status") or "").strip(),
            "auth_start": (row.findtext("Start_Date") or "").strip()[:10],
            "auth_end": (row.findtext("End_Date") or "").strip()[:10],
            "member_id": (row.findtext("Insurance_Policy_No") or "").strip(),
            "mco": (row.findtext("Payor") or "").strip(),
            "billing_diagnosis": (row.findtext("Billing_Diagnosis") or "").strip(),
        }

    rows: list[dict] = []
    for row in billing_root.findall(".//Billing_Summary_View"):
        bs_id = (row.findtext("Billing_Summary_ID") or "").strip()
        doc_date = (row.findtext("Document_Date") or "").strip()
        if not bs_id or (doc_date and doc_date < cutoff):
            continue

        auth_id = (
            row.findtext("AuthID__x0028_for_joining_x0029_")
            or ""
        ).strip()
        auth = auth_by_join.get(auth_id, {})
        rows.append({
            "bs_id": bs_id,
            "patient_account_number": (row.findtext("STBID") or "").strip(),
            "lauris_id": (row.findtext("Key") or "").strip(),
            "key": (row.findtext("Key") or "").strip(),
            "name": (row.findtext("Name") or "").strip(),
            "auth_id": auth_id,
            "doc_date": doc_date[:10],
            "member_id": auth.get("member_id", ""),
            "mco": auth.get("mco", ""),
            "auth_number": auth.get("auth_number", ""),
            "auth_status": auth.get("auth_status", ""),
            "auth_start": auth.get("auth_start", ""),
            "auth_end": auth.get("auth_end", ""),
            "billing_diagnosis": auth.get("billing_diagnosis", ""),
            "service": (row.findtext("Service_Name") or "").strip(),
            "billing_amount": float(row.findtext("Billing_Amount") or "0"),
            "billing_units": (row.findtext("Billing_Units") or "").strip(),
            "modifier": (row.findtext("Billing_Modifier") or "").strip(),
        })

    logger.info("Claim member bridge rows loaded", count=len(rows))
    return rows


def enrich_with_claimmd_notes(
    claims: List[dict],
    notes_by_member_dos: Dict[str, List[str]],
    denials_by_member_dos: Dict[str, List[str]],
) -> None:
    """
    Enrich outstanding claims with Claim.MD notes and denial
    reasons, matched by member_id + DOS.
    Modifies claims in place.
    """
    linked_notes = 0
    linked_denials = 0
    for item in claims:
        member = item.get("member_id", "")
        dos = item.get("doc_date", "")
        if not member or not dos:
            continue
        key = f"{member}_{dos}"
        if key in notes_by_member_dos:
            item["claimmd_notes"] = " || ".join(
                notes_by_member_dos[key]
            )[:500]
            linked_notes += 1
        if key in denials_by_member_dos:
            item["claimmd_denials"] = " || ".join(
                denials_by_member_dos[key]
            )[:500]
            linked_denials += 1

    logger.info(
        "Claim.MD data linked",
        notes=linked_notes,
        denials=linked_denials,
    )


async def fetch_claimmd_notes_and_denials():
    """
    Fetch all Claim.MD notes and denial reasons.
    Returns (notes_by_member_dos, denials_by_member_dos).
    """
    from sources.claimmd_api import ClaimMDAPI

    api = ClaimMDAPI()

    # Get ALL notes
    result = await api._post("notes", {"ClaimID": ""})
    all_notes = result.get("notes", [])
    logger.info("Claim.MD notes fetched", count=len(all_notes))

    # Get responses for member_id mapping + denial reasons
    raw = await api.get_claim_responses(
        response_id="1700000000"
    )

    # Build claimmd_id -> member+DOS mapping
    cid_to_member: Dict[str, dict] = {}
    denials_by_member_dos: Dict[str, List[str]] = defaultdict(
        list
    )
    for r in raw:
        cid = str(r.get("claimmd_id", ""))
        member = (r.get("ins_number", "") or "").strip()
        dos = (r.get("fdos", "") or "").strip()[:10]
        status = r.get("status", "")
        if cid and member:
            cid_to_member[cid] = {
                "member": member,
                "dos": dos,
            }
        # Capture ALL denial + warning messages (not just first)
        if member and dos:
            for m in r.get("messages", []) or []:
                if isinstance(m, dict):
                    msg_status = m.get("status", "")
                    msg_text = m.get("message", "")[:200]
                    # Include R (rejection) and W (warning) messages
                    if msg_text and msg_status in ("R", "W"):
                        denials_by_member_dos[
                            f"{member}_{dos}"
                        ].append(msg_text)

    # Build notes by member+DOS
    notes_by_member_dos: Dict[str, List[str]] = defaultdict(
        list
    )
    for n in all_notes:
        cid = str(n.get("claimmd_id", ""))
        if cid in cid_to_member:
            key = (
                f"{cid_to_member[cid]['member']}_"
                f"{cid_to_member[cid]['dos']}"
            )
            note_text = n.get("note", "").strip()
            user = n.get("username", "").strip()
            dt = n.get("date_time", "")[:16]
            if note_text:
                notes_by_member_dos[key].append(
                    f"[{dt}] {user}: {note_text}"
                )

    logger.info(
        "Claim.MD lookups built",
        notes_keys=len(notes_by_member_dos),
        denial_keys=len(denials_by_member_dos),
    )
    return dict(notes_by_member_dos), dict(
        denials_by_member_dos
    )


def is_claim_in_ar(
    member_id: str,
    dos: Optional[date],
    ar_claims: List[dict],
) -> Optional[dict]:
    """
    Check if a claim (by member_id + DOS) appears in the AR
    outstanding claims list.
    Returns the matching AR entry or None.
    """
    if not member_id or not dos:
        return None
    dos_str = dos.isoformat()[:10] if isinstance(dos, date) else str(dos)[:10]
    for ar in ar_claims:
        if (
            ar.get("member_id", "") == member_id
            and ar.get("doc_date", "") == dos_str
        ):
            return ar
    return None


def generate_unified_excel(
    claims: List[dict],
    output_path: str | None = None,
) -> str:
    """Generate the unified outstanding claims Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment,
    )
    from openpyxl.utils import get_column_letter

    if output_path is None:
        output_path = str(
            unique_report_path(
                "Unified Outstanding Claims",
                "outstanding_claims_unified",
                ".xlsx",
            )
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Outstanding Claims"
    headers = [
        "Client Name", "Unique ID", "Service", "DOS",
        "Entity", "Billing Amt", "Received", "Outstanding",
        "AR Status", "MCO", "Member ID", "Diagnosis",
        "Auth #", "Auth Start", "Auth End", "Auth Status",
        "Invoice Date", "Claim.MD Notes", "Denial Reason",
    ]
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(
        start_color="1F4E79", end_color="1F4E79",
        fill_type="solid",
    )
    due_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6",
        fill_type="solid",
    )
    under_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC",
        fill_type="solid",
    )
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf
        c.fill = hfill
        c.border = tb

    for ri, item in enumerate(
        sorted(claims, key=lambda x: x.get("name", "")), 2
    ):
        vals = [
            item.get("name", ""),
            item.get("key", ""),
            item.get("service", ""),
            item.get("doc_date", ""),
            item.get("region", ""),
            item.get("billing_amount", 0),
            item.get("total_received", 0),
            item.get("outstanding", 0),
            item.get("ar_status", ""),
            item.get("mco", ""),
            item.get("member_id", ""),
            item.get("billing_diagnosis", "")[:50],
            item.get("auth_number", ""),
            item.get("auth_start", ""),
            item.get("auth_end", ""),
            item.get("auth_status", ""),
            item.get("invoice_date", ""),
            item.get("claimmd_notes", "")[:500],
            item.get("claimmd_denials", "")[:500],
        ]
        fill = (
            due_fill
            if item.get("ar_status") == "Due"
            else under_fill
        )
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = tb
            cell.fill = fill

    widths = [
        25, 12, 8, 12, 28, 12, 12, 12, 14, 20, 18, 35,
        16, 12, 12, 12, 12, 60, 60,
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}"
        f"{len(claims) + 1}"
    )
    wb.save(output_path)
    sync_report_file(Path(output_path), "Unified Outstanding Claims")
    logger.info(
        "Unified report saved",
        path=output_path,
        claims=len(claims),
    )
    return output_path
