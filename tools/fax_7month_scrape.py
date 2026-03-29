#!/usr/bin/env python3
"""
7-Month Fax Scrape with Hourly Excel Reports
=============================================
Scrapes all 4 fax systems going back 7 months.
Skips entries already in the database.
Generates an Excel report every hour so progress can be monitored.
"""
import asyncio
import os
import sys
import time
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ["DRY_RUN"] = "false"

from dotenv import load_dotenv
load_dotenv(str(Path(__file__).resolve().parent.parent / ".env"))

import actions.fax_tracker as ft
ft.MAX_PDF_DOWNLOADS_PER_RUN = 9999  # No cap

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "claims_history.db"
REPORT_PATH = Path(__file__).resolve().parent.parent / "docs" / "fax_tracking_report.xlsx"


def generate_report():
    """Generate Excel report with PDF scan timestamp column."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.row_factory = sqlite3.Row

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sf = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rf = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    denied_f = Font(color="FF0000", bold=True)
    approved_f = Font(color="008000", bold=True)
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "FAX TRACKING REPORT — 7 Month Scrape"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    for i, h in enumerate(["Source", "Total", "With Client Name", "Approved", "Denied"], 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = hf; c.fill = hfill; c.border = tb
    sources = conn.execute(
        "SELECT source, COUNT(*) as cnt FROM fax_log GROUP BY source ORDER BY source"
    ).fetchall()
    row = 5
    for s in sources:
        src = s["source"]
        ws.cell(row=row, column=1, value=src).border = tb
        ws.cell(row=row, column=2, value=s["cnt"]).border = tb
        ws.cell(row=row, column=3, value=conn.execute(
            "SELECT COUNT(*) FROM fax_log WHERE source=? AND client_name != '' AND client_name IS NOT NULL",
            (src,),
        ).fetchone()[0]).border = tb
        ws.cell(row=row, column=4, value=conn.execute(
            "SELECT COUNT(*) FROM fax_log WHERE source=? AND notes LIKE '%approved%'", (src,),
        ).fetchone()[0]).border = tb
        ws.cell(row=row, column=5, value=conn.execute(
            "SELECT COUNT(*) FROM fax_log WHERE source=? AND notes LIKE '%denied%'", (src,),
        ).fetchone()[0]).border = tb
        row += 1
    total_all = conn.execute("SELECT COUNT(*) FROM fax_log").fetchone()[0]
    total_named = conn.execute(
        "SELECT COUNT(*) FROM fax_log WHERE client_name != '' AND client_name IS NOT NULL"
    ).fetchone()[0]
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_all).font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_named).font = Font(bold=True)
    for w, width in enumerate([32, 12, 18, 12, 12], 1):
        ws.column_dimensions[get_column_letter(w)].width = width

    # Data sheets with PDF Scan Timestamp column
    def make_sheet(title, source):
        cols = [
            "Date", "Client Name", "Entity", "MCO", "Fax #",
            "Auth Status", "Service", "Diagnosis", "Auth #",
            "Auth Period", "Doc Type", "Source", "PDF Scan Completed",
        ]
        widths = [22, 25, 30, 24, 14, 13, 10, 22, 16, 26, 12, 28, 22]
        ws2 = wb.create_sheet(title=title)
        for i, col in enumerate(cols, 1):
            c = ws2.cell(row=1, column=i, value=col)
            c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal="center"); c.border = tb
        rows = conn.execute(
            "SELECT * FROM fax_log WHERE source=? ORDER BY fax_date DESC", (source,),
        ).fetchall()
        for ri, r in enumerate(rows, 2):
            notes = r["notes"] or ""
            svc = notes.split("Service: ")[1].split(" |")[0] if "Service: " in notes else ""
            dx = notes.split("Dx: ")[1].split(" |")[0] if "Dx: " in notes else ""
            auth_st = r["auth_status"] or ""
            if "denied" in notes.lower():
                ds = "DENIED"
            elif "approved" in notes.lower():
                ds = "APPROVED"
            elif auth_st == "submitted":
                ds = "Sent"
            elif "success" in auth_st:
                ds = "Success"
            else:
                ds = auth_st
            # PDF scan timestamp from reviewed_at
            scan_ts = r["reviewed_at"] or ""
            vals = [
                r["fax_date"] or "", r["client_name"] or "", r["company"] or "",
                r["mco"] or "", r["fax_number"] or "", ds, svc, dx,
                r["auth_number"] or "", r["auth_dates"] or "",
                r["document_type"] or "", r["source"] or "", scan_ts,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.border = tb
                if val == "DENIED":
                    cell.font = denied_f
                elif val == "APPROVED":
                    cell.font = approved_f
                elif "received" in source and ri % 2 == 0:
                    cell.fill = rf
                elif "received" not in source and ri % 2 == 0:
                    cell.fill = sf
        for i, w in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

    make_sheet("Lauris Sent", "lauris_sent")
    make_sheet("Nextiva nmoyern Sent", "nextiva_nmoyern_sent")
    make_sheet("Nextiva nmoyern2 Sent", "nextiva_nmoyern2_sent")
    make_sheet("Nextiva Received", "nextiva_nmoyern_received")

    conn.close()
    wb.save(str(REPORT_PATH))
    print(f"[{datetime.now().strftime('%I:%M %p')}] Report saved: {REPORT_PATH.name} | {total_all} entries | {total_named} named")


async def run_scrape():
    from actions.fax_tracker import (
        scrape_lauris_fax_history,
        scrape_nextiva_sent_faxes,
        scrape_nextiva_received_faxes,
    )
    from lauris.billing import LaurisSession

    start = date(2025, 8, 25)  # Fixed start date per user request
    t_total = time.time()
    last_report = time.time()

    def maybe_report():
        nonlocal last_report
        if time.time() - last_report >= 3600:  # 1 hour
            generate_report()
            last_report = time.time()

    print(f"{'='*70}")
    print(f"  7-MONTH FAX SCRAPE")
    print(f"  Date range: {start} to {date.today()}")
    print(f"  Existing entries will be skipped")
    print(f"  Report generated every hour")
    print(f"{'='*70}")

    # Initial report
    generate_report()

    # System 1: Lauris
    print(f"\n[{datetime.now().strftime('%I:%M %p')}] === SYSTEM 1: LAURIS ===")
    t0 = time.time()
    async with LaurisSession(headless=True) as lauris:
        r1 = await scrape_lauris_fax_history(lauris, start)
    named1 = sum(1 for r in r1 if r.get("client_name"))
    print(f"  New: {len(r1)} | Named: {named1}/{len(r1)} | Time: {time.time()-t0:.0f}s")
    maybe_report()

    # System 2: Nextiva nmoyern sent
    print(f"\n[{datetime.now().strftime('%I:%M %p')}] === SYSTEM 2: NEXTIVA nmoyern SENT ===")
    t0 = time.time()
    r2 = await scrape_nextiva_sent_faxes("nmoyern", full_scrape=True)
    named2 = sum(1 for r in r2 if r.get("client_name"))
    print(f"  New: {len(r2)} | Named: {named2}/{len(r2)} | Time: {time.time()-t0:.0f}s")
    maybe_report()

    # System 3: Nextiva nmoyern2 sent
    print(f"\n[{datetime.now().strftime('%I:%M %p')}] === SYSTEM 3: NEXTIVA nmoyern2 SENT ===")
    t0 = time.time()
    r3 = await scrape_nextiva_sent_faxes("nmoyern2", full_scrape=True)
    named3 = sum(1 for r in r3 if r.get("client_name"))
    print(f"  New: {len(r3)} | Named: {named3}/{len(r3)} | Time: {time.time()-t0:.0f}s")
    maybe_report()

    # System 4: Nextiva received
    print(f"\n[{datetime.now().strftime('%I:%M %p')}] === SYSTEM 4: NEXTIVA RECEIVED ===")
    t0 = time.time()
    r4 = await scrape_nextiva_received_faxes(full_scrape=True)
    named4 = sum(1 for r in r4 if r.get("client_name"))
    print(f"  New: {len(r4)} | Named: {named4}/{len(r4)} | Time: {time.time()-t0:.0f}s")

    total_new = len(r1) + len(r2) + len(r3) + len(r4)
    total_named = named1 + named2 + named3 + named4
    elapsed = time.time() - t_total

    print(f"\n{'='*70}")
    print(f"  COMPLETE")
    print(f"  New entries: {total_new} | Named: {total_named}")
    print(f"  Total time: {elapsed/60:.0f} minutes")
    print(f"{'='*70}")

    # Final report
    generate_report()


if __name__ == "__main__":
    asyncio.run(run_scrape())
